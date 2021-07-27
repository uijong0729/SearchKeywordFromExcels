import datetime
from openpyxl import Workbook
wb = Workbook()
ws = wb.active

# 날짜
ws["A1"] = datetime.datetime.today()

# 합계 계산
ws["A2"] = "=SUM(1,2,3)"

# 평균 계산
ws["A3"] = "=AVERAGE(1,2,3)"

# 수식에서의 셀의 참조
ws["A4"] = 10
ws["A5"] = 20
ws["A6"] = "=SUM(A4:A5)"

wb.save("sample_formula.xlsx")
wb.close()
from openpyxl import load_workbook

# ws.values => 셀 객체가 아니라 value정보를 바로 가져온다. 셀이 수식인 경우 수식 그대로 가져 옴.
wb = load_workbook("sample_formula.xlsx")
ws = wb.active
for row in ws.values:
    for cell in row:
        print(cell)
wb.close()

# data_only=True => 셀이 수식인 경우 계산된 결과를 가져옴.
wb = load_workbook("sample_formula.xlsx", data_only=True)
ws = wb.active
for row in ws.values:
    for cell in row:
        # 계산 되지 않은 상태의 데이터는 None이라고 표시 => 파일을 열었다 저장하고 끄면 해결 됨
        print(cell)

wb.save("sample_formula_2.xlsx")
wb.close()
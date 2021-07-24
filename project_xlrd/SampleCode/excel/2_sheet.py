from openpyxl import Workbook

wb = Workbook()

ws = wb.create_sheet()                      # 새로운 시트를 기본이름으로 생성
ws.title = "MySheet"
ws.sheet_properties.tabcolor = "ff66ff"

ws1 = wb.create_sheet("YourSheet")     # 새로운 시트를 지정이름으로 생성
ws2 = wb.create_sheet("MySheet1", 2)     # 새로운 시트를 2번째 인덱스에 지정이름으로 생성
ws3 = wb["YourSheet"] #Dict 형태로 sheet에 접근
print(wb.sheetnames)

# Sheet의 데이터 복사
ws3["A1"] = "Test"
target = wb.copy_worksheet(ws3)
target.title = "copied sheet"

wb.save("sample.xlsx")
wb.close()

from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws.title = "MySheet"

# A1 셀에 1이라는 값을 입력
ws["A1"] = 1
ws["A2"] = 2
ws["A3"] = 3
ws["B1"] = 4
ws["B2"] = 5
ws["B3"] = 6
c = ws.cell(column=3, row=1, value=10) # ws["C1"].value = 10 과 동일한 동작 
print(c.value) # A1셀 

print(ws["A1"])         # A1셀의 정보를 출력
print(ws["A1"].value)   # A1셀의 값을 출력
print(ws["Z99"].value)   # 값이 없을 때는 None을 출력

# row = 1, 2, 3, ...
# column = A, B, C, ...
print(ws.cell(row=1, column=1).value) # A1셀 

###################################################################################
# 반복문을 이용해서 랜덤 숫자 채우기
###################################################################################
from random import *

data = 1
for x in range(1, 11):
    for y in range(1, 11):
        # ws.cell(row=x, column=y, value=randint(0, 100))
        ws.cell(row=x, column=y, value=data)
        data += 1 # 파이썬은 증감연산자가 없음


###################################################################################
# 셀을 영역 단위로 취급하기
###################################################################################
wb = Workbook()
ws = wb.active

# 1줄씩 데이터 넣기
# append에 리스트 형식으로 인수를 넣어준다.
ws.append(["번호", "영어", "수학"])
for i in range(1, 11):
    ws.append([i, randint(0, 100), randint(0, 100)])

# 영어 Column만 가져오기
col_B = ws["B"]
print(col_B)
for cell in col_B:
    print(cell.value)

# 영어, 수학 Column을 함께 가져오기
col_range = ws["B:C"] 
for cols in col_range:
    for cell2 in cols:
        print(cell2.value)

# 1번째 row만 가져오기
row_title = ws[1]
for rCell in row_title:
    print(rCell.value)

# 2번째 줄에서 6번째 줄까지 가지고 오기
row_title = ws[2:6]
for rows in row_title:
    for cell in rows:
        print(cell.value, end=" ")
    print()

# 2번째 줄부터 마지막 줄 까지
row_range = ws[2:ws.max_row]
for rows in row_range:
    for cell in rows:
        print(cell.value, end=" ")
    print()

# 셀의 좌표정보 가져오기 
from openpyxl.utils.cell import coordinate_from_string
row_range = ws[2:ws.max_row]
for rows in row_range:
    for cell in rows:
        # A1과 같은 형식으로 좌표정보 취득
        print(cell.coordinate, end=" ")

        # ('A', 2)와 같은 튜플 형식으로 좌표정보 취득
        # xy[0] == 'A'
        # xy[1] == 1
        xy = coordinate_from_string(cell.coordinate)
        print(xy, end=" ")
        print(xy[0], end=" ")
        print(xy[1], end=" ")

    print()

# 튜플과 리스트의 차이 
# 리스트는 [ ]으로 둘러싸지만 튜플은 ( )으로 둘러싼다.
# 리스트는 그 값의 생성, 삭제, 수정이 가능하지만 튜플은 그 값을 바꿀 수 없다.

# 모든 줄 가져오기1
print(tuple(ws.rows))
for row in tuple(ws.rows):
    # 행의 두 번째 열의 값을 출력
    print(row[1].value)

# 모든 줄 가져오기2
for row in ws.iter_rows():
    print(row)

# 모든 컬럼 가져오기1
print(tuple(ws.columns))
for column in tuple(ws.columns):
    # 열의 첫 번째 행의 값을 출력
    print(column[0].value)

# 모든 컬럼 가져오기2
for column in ws.iter_cols():
    print(column[0].value)

# 모든 행에서 가져올 행범위 지정
for row in ws.iter_rows(min_row=1, max_row=5):
    print(row[2].value)

# 모든 행에서 가져올 행, 열범위 지정
for row in ws.iter_rows(min_row=1, max_row=5, min_col=1, max_col=3):
    print(row[2].value, row[1].value)

wb.save("sample.xlsx")
wb.close()
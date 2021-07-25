# 파일 불러오기
from openpyxl import load_workbook

wb = load_workbook("sample.xlsx")

# 활성화 된 시트
ws = wb.active

# Cell 데이터 불러오기
for x in range(1, 11):
    for y in range(1, 11):
        # end = " " 데이터마다 띄어쓰기가 삽입된다.
        print(ws.cell(row=x, column=y).value, end=" ")
    # 줄 바꿈
    print()

# cell 개수를 모를 때
for x in range(1, ws.max_row + 1):
    for y in range(1, ws.max_column + 1):
        print(ws.cell(row=x, column=y).value, end=" ")
    print()

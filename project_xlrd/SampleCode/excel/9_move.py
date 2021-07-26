from openpyxl import load_workbook
wb = load_workbook("sample.xlsx")
ws = wb.active

# 번호 영어 수학
# 번호 [국어] 영어 수학

# B1:C11 범위의 셀들을 오른쪽으로 1열 옮김
# 음수 값을 지정하면 왼쪽으로 옮김
ws.move_range("B1:C11", rows=0, cols=1)

ws["B1"] = "국어"

wb.save("sample_korean.xlsx")
wb.close()